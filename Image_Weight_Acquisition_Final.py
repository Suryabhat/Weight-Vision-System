import cv2
from picamera2 import Picamera2
import os
from datetime import datetime
import RPi.GPIO as GPIO
import time
import openpyxl
from hx711 import HX711

image_folder = "Arecanut_images"
os.makedirs(image_folder, exist_ok=True)

image_counter = 1

def adjust_lens_position():
    
    lens_position = 8.0
    print(f"Adjusting lens position to {lens_position}...")
    
def capture_image_and_measure_weight(picam2, hx, sheet):
    global image_counter

    # Check if the sheet is not empty
    if sheet.max_row > 1:
        # Read the last filled row from the Excel sheet
        last_row = sheet.max_row

        # Extract the image counter value from the 'C' column in the last row
        last_image_counter_str = sheet.cell(row=last_row, column=3).value.split('_')[1]
        
        # Convert the counter value to an integer
        last_image_counter = int(last_image_counter_str)

        # Increment the image counter for the next image
        image_counter = last_image_counter + 1

    im = picam2.capture_array()
    cv2.imshow("Camera", im)

    key = cv2.waitKey(1)
    if key == ord('q'):
        return True  # Break out of the loop
    elif key == ord(' '):  # Capture image and measure weight when the space key is pressed
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        image_path = os.path.join(image_folder, f"image_{image_counter}_{timestamp}.jpg")
        picam2.capture_file(image_path)
        print(f"Image captured and saved as {image_path}")

        # Measure weight using HX711
        weight = hx.get_weight_mean()
        timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
        print(f'Timestamp: {timestamp}, Weight: {weight} grams, Image: {image_path}')

        # Store data in the Excel sheet
        row = (timestamp, weight, image_path)
        sheet.append(row)

        # Update the image counter value in the Excel sheet
        sheet.cell(row=sheet.max_row, column=3, value=f"image_{image_counter}_{timestamp}.jpg")

        return False  # Continue the loop

    return False

def setup_hx711():
    hx = HX711(dout_pin=10, pd_sck_pin=11)
    err = hx.zero()
    if err:
        raise ValueError('Tare is unsuccessful.')

    while True:
        try:
            known_weight_grams = input('Put known weight on the scale and then press Enter: ')
            if known_weight_grams:
                ratio = hx.get_data_mean() / float(known_weight_grams)
                hx.set_scale_ratio(ratio)
                break
            else:
                print('Please enter a valid weight.')

        except ValueError:
            print('Invalid input. Please enter a valid weight.')

    return hx

def setup_excel():
    if os.path.exists('weight_data.xlsx'):
        # If the file exists, load the workbook
        wb = openpyxl.load_workbook('weight_data.xlsx')
        sheet = wb.active
    else:
        # If the file doesn't exist, create a new workbook and sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = 'Timestamp'
        sheet['B1'] = 'Weight (grams)'
        sheet['C1'] = 'Image Path'

    return wb, sheet

if __name__ == "__main__":
    try:
        GPIO.setmode(GPIO.BCM)

        # Initialize PiCamera
        picam2 = Picamera2()
        picam2.preview_configuration.main.size = (1280, 720)
        picam2.preview_configuration.main.format = "RGB888"
        picam2.preview_configuration.align()
        picam2.configure("preview")
        picam2.start()
        
        # Initialize HX711
        hx_instance = setup_hx711()

        # Set up Excel workbook and sheet
        wb, sheet = setup_excel()

        # Main loop
        while True:
            if capture_image_and_measure_weight(picam2, hx_instance, sheet):
                break

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        GPIO.cleanup()
        wb.save('weight_data.xlsx')
        cv2.destroyAllWindows()
        wb.close()

